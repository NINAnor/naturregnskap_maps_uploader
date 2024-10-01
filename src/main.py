import logging
import pathlib
import re

import click
import environ
from api import get_client
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import load_workbook
from upload import check_map, create_layer, create_project_folder
from yaml import safe_load

env = environ.Env()
BASE_DIR = pathlib.Path(__file__).parent.parent
environ.Env.read_env(str(BASE_DIR / ".env"))

DEBUG = env.bool("DEBUG", default=False)

TITILER_URL = env("TITILER_URL")

logging.basicConfig(level=(logging.DEBUG if DEBUG else logging.INFO))


template_env = Environment(
    loader=FileSystemLoader(pathlib.Path(__file__).parent / "templates"),
    autoescape=select_autoescape(),
)


def str_to_snake_case(text: str) -> str:
    """
    Convert to snake_case.
    - Header One -> header_one
    - HeaderTwo -> header_two
    - headerTwo -> header_two
    - HEADER -> header
    - EPSG Code -> epsg_code
    - someURLValue -> some_url_value
    """
    text = text.strip()
    #  uppercase (>2 letters) converted to lowercase
    text = re.sub(r"\b[A-Z]{2,}\b", lambda m: m.group(0).lower(), text)
    # to snake_case
    text = re.sub(r"(?<!^)(?<![A-Z])(?=[A-Z])", "_", text).replace(" ", "_").lower()
    return text


def load_workbook_sheet(wb, sheet_name: str, skip_first_row: bool = True):
    sheet = wb[sheet_name]
    rows = sheet.iter_rows()

    # first row is header
    if skip_first_row:
        next(rows)
        header = [cell.value.strip() for cell in next(rows) if cell.value]
        return rows, header

    # first row is data, header is None
    return rows, None


@click.command()
@click.argument("url")
@click.argument("map_slug")
@click.option("--schema", default="schema.xlsx", type=click.Path())
@click.option("--style", default="style.yml", type=click.Path())
@click.option(
    "--wd",
    default=".",
    type=click.Path(dir_okay=True, exists=True, readable=True, path_type=pathlib.Path),
)
def start(url: str, map_slug: str, schema: str, style: str, wd: pathlib.Path) -> None:
    style_path = wd / style
    schema_path = wd / schema

    style_index = {}
    with style_path.open("r") as style_file:
        conf = safe_load(style_file)
        style_index = conf["datasets"]

    wb = load_workbook(str(schema_path))

    # Load layer group metadata (projectMetadata)
    rows, header = load_workbook_sheet(wb, "projectMetadata", True)
    project_metadata = {}
    for row in rows:
        row = [
            cell.value.strip() if isinstance(cell.value, str) else cell.value
            for cell in row
        ]
        if not any(row):
            continue
        project_metadata = dict(zip(header, row, strict=True))

        break

    logging.debug("found_project %s" % project_metadata)

    # Load project contacts
    project_metadata["contacts"] = []
    rows, header = load_workbook_sheet(wb, "projectContacts", True)
    for row in rows:
        row = [
            cell.value.strip() if isinstance(cell.value, str) else cell.value
            for cell in row
        ]
        if not any(row):
            continue
        project_metadata["contacts"].append(dict(zip(header, row, strict=True)))

    logging.info("project_metadata %s" % project_metadata)

    # (re)create project layer group in django
    TOKEN = env("AUTH_TOKEN")
    logging.debug(f"using TOKEN: {TOKEN}")
    client = get_client(base_url=url, token=TOKEN)
    check_map(client, map_slug)
    project_slug = create_project_folder(
        client,
        map_slug,
        project_metadata,
        template_env,
    )

    # TODO: read and save also project owner and contributors

    # Load layer metadata (datasetMetadata) into nested dict
    # lyr_metadata = {h1: {h2: value}}
    # h1 = first row, filled with previous value if empty
    # h2 = second row, should not contain empty values
    rows, header = load_workbook_sheet(wb, "datasetMetadata", False)
    h1, last_value = [], None
    h1 = [
        (last_value := str_to_snake_case(cell.value) if cell.value else last_value)
        for cell in next(rows)
    ]
    h2 = [cell.value.strip() for cell in next(rows) if cell.value]
    h2_snake_case = list(map(str_to_snake_case, h2))

    lyr_metadata = {h1[0]: {}}
    for row in rows:
        row = [
            cell.value.strip() if isinstance(cell.value, str) else cell.value
            for cell in row
        ]
        if not any(row):
            continue
        dataset_metadata = dict(zip(h2, row, strict=True))
        dataset_metadata = {
            k: (v if v is not None else "") for k, v in dataset_metadata.items()
        }
        for h1_key, h2_key, value in zip(h1, h2_snake_case, row, strict=False):
            if h1_key not in lyr_metadata:
                lyr_metadata[h1_key] = {}
            lyr_metadata[h1_key][h2_key] = value if value is not None else ""

        logging.debug(f"lyr_metadata: {lyr_metadata}")
        logging.debug(f'uploading {dataset_metadata["datasetAlias"]}')

        layer_slug = f"{project_slug}_{dataset_metadata['datasetName']}"

        if dataset_metadata["skip"] == 0:
            layer_style = (
                style_index[layer_slug]
                if layer_slug in style_index
                else style_index["DEFAULT"]
            )

            create_layer(
                client=client,
                map_slug=map_slug,
                layer=dataset_metadata,
                project=project_slug,
                slug=layer_slug,
                style=layer_style,
                wd=wd,
                titiler_url=TITILER_URL,
                template_env=template_env,
                lyr_metadata=lyr_metadata,
                project_metdata=project_metadata,
            )


if __name__ == "__main__":
    start()
